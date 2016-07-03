#!/usr/bin/env python3

"""xml2xlsx -- XMLish API for OpenPyXL

Provided XML Element Tree will return openpyxl.Workbook object.

This XMLish API tries to mimic OpenPyXL API and is supposed to be thin wrapper
on it.
"""

import sys
import tempfile

from lxml import etree
import openpyxl

# TODO: check if font and fill names are distinct
# TODO: check if font and fill names exists when applying
# TODO: check if cells are not being overwritten (how?)
# TODO: main() should print file name


def create_cell(sheet, cell_element):
    if 'pos' in cell_element.attrib:
        return sheet[cell_element.attrib['pos']]
    return sheet.cell(
        row=int(cell_element.attrib['row']),
        column=int(cell_element.attrib['column']),
    )


def process(root):
    fonts = {}
    fills = {}
    workbook = openpyxl.Workbook()

    # get fonts and fills
    for f in root.xpath('font'):
        fonts[f.text] = openpyxl.styles.Font(**dict(f.attrib))
    for pf in root.xpath('fill | pattern_fill'):
        fills[pf.text] = openpyxl.styles.PatternFill(**dict(pf.attrib))

    for i, s in enumerate(root.xpath('sheet')):
        # First sheet is already there and does not have to be created
        if i == 0:
            sheet = workbook.active
            sheet.title = s.attrib['title']
        else:
            sheet = workbook.create_sheet(title=s.attrib['title'])

        # add freeze for sheet (if defined)
        for fp in s.xpath('freeze | freeze_pane | freeze_panes'):
            sheet.freeze_panes = create_cell(sheet, cell_element=fp)
            break

        # cells can be defined both as <c> or <cell>
        for c in s.xpath('c | cell'):
            cell = create_cell(sheet, cell_element=c)

            try:
                cell.value = float(c.text)
            except ValueError:
                cell.value = c.text

            if 'font' in c.attrib:
                cell.font = fonts[c.attrib['font']]
            if 'fill' in c.attrib:
                cell.fill = fills[c.attrib['fill']]

    return workbook


def main():
    workbook = process(etree.fromstring(sys.stdin.read()))
    with tempfile.SpooledTemporaryFile(max_size=32*1024*1024) as f:
        workbook.save(f)
        sys.stdout.buffer.write(f.read())

if __name__ == "__main__":
    main()
